import os
import glob
from pathlib import Path
from xml.etree import ElementTree as ET
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

# === Detección flexible de BASE_DIR ===
# Si existe ./facturas_xml, úsalo; si no, y estás dentro de facturas_xml, usa "."
if Path("./facturas_xml").exists():
    BASE_DIR = Path("./facturas_xml")
else:
    BASE_DIR = Path(".")

OUTPUT   = "reporte_retenciones_{anio}.xlsx"
MESES = [f"{i:02d}" for i in range(1, 13)]

def txt(el, path):
    if el is None:
        return ""
    node = el.find(path)
    return (node.text or "").strip() if node is not None else ""

def try_parse_xml(raw: str):
    try:
        return ET.fromstring(raw)
    except ET.ParseError:
        return ET.fromstring(raw.strip())

def parse_autorizacion(root: ET.Element):
    numero_aut = txt(root, "./numeroAutorizacion")
    fecha_aut  = txt(root, "./fechaAutorizacion")
    comp_cdata = txt(root, "./comprobante")
    return numero_aut, fecha_aut, comp_cdata

def norm_fecha_ddmmyyyy_to_iso(s):
    if not s:
        return ""
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except:
            pass
    return s

def to_float(s):
    if s is None:
        return 0.0
    s = str(s).replace(",", ".").strip()
    try:
        return float(s)
    except:
        return 0.0

def parse_one_xml(xml_path: Path):
    """
    Devuelve:
      header_info: {'razon_suj','ruc_suj'}
      row: dict con UNA FILA por comprobante de retención:
           - BASE IMPONIBLE = suma de líneas
           - VALOR RETENIDO = suma de líneas
           - % DE RETENCION = PRIMER porcentajeRetener encontrado
    """
    raw = xml_path.read_text(encoding="utf-8", errors="ignore")
    root = try_parse_xml(raw)
    tag = root.tag.lower()
    if "}" in tag:
        tag = tag.split("}", 1)[1]

    numero_aut = ""
    comp_root  = None

    if tag == "autorizacion":
        numero_aut, _, comp_cdata = parse_autorizacion(root)
        comp_root = try_parse_xml(comp_cdata)
    else:
        comp_root = root  # comprobante sin envoltorio

    tag_comp = comp_root.tag.lower()
    if "}" in tag_comp:
        tag_comp = tag_comp.split("}", 1)[1]

    header = {"razon_suj": "", "ruc_suj": ""}

    if tag_comp not in ("comprobanteretencion", "comprobanteretención"):
        return header, None

    infoTrib = comp_root.find("./infoTributaria")
    infoRet  = comp_root.find("./infoCompRetencion")

    estab = txt(infoTrib, "estab")
    pto   = txt(infoTrib, "ptoEmi")
    sec   = txt(infoTrib, "secuencial")
    serie = f"{estab}-{pto}-{sec}"

    # Autorización: si no hubo numeroAut, intenta claveAcceso
    if not numero_aut:
        numero_aut = txt(infoTrib, "claveAcceso")

    fecha_emision = norm_fecha_ddmmyyyy_to_iso(
        txt(infoRet, "fechaEmision") or txt(infoRet, "fechaEmisionDocSustento")
    )
    ruc_agente = txt(infoTrib, "ruc")

    # Encabezado del contribuyente (sujeto retenido)
    header["razon_suj"] = txt(infoRet, "razonSocialSujetoRetenido")
    header["ruc_suj"]   = txt(infoRet, "identificacionSujetoRetenido")

    # ——— Capturar líneas (v2.0.0 y v1.0.0) ———
    lines = comp_root.findall(".//retencion")
    if not lines:
        lines = comp_root.findall(".//impuestos/impuesto")

    base_total = 0.0
    valor_retenido_total = 0.0

    porcentaje_retencion = ""  # tomaremos SOLO el primero encontrado
    for ln in lines:
        base_total += to_float(txt(ln, "baseImponible"))
        valor_retenido_total += to_float(txt(ln, "valorRetenido"))
        if porcentaje_retencion == "":  # primer porcentaje que aparezca
            p = txt(ln, "porcentajeRetener")
            if p:
                try:
                    porcentaje_retencion = float(str(p).replace(",", ".").strip())
                except:
                    porcentaje_retencion = p  # deja texto tal cual si no parsea

    row = {
        "N° DE COMPROBANTES DE RETENCION": serie,
        "N° DE AUTORIZACION DEL COMPROBANTE": numero_aut,
        "FECHA DE EMISION DEL COMPROBANTE DE RETENCION": fecha_emision,
        "RUC DEL AGENTE DE RETENCION": ruc_agente,
        "BASE IMPONIBLE": base_total,
        "%  DE RETENCION": porcentaje_retencion,
        "VALOR RETENIDO": valor_retenido_total,
    }
    return header, row

def write_excel(anio, rows, header_razon, header_ruc, out_path: Path):
    df = pd.DataFrame(rows, columns=[
        "N° DE COMPROBANTES DE RETENCION",
        "N° DE AUTORIZACION DEL COMPROBANTE",
        "FECHA DE EMISION DEL COMPROBANTE DE RETENCION",
        "RUC DEL AGENTE DE RETENCION",
        "BASE IMPONIBLE",
        "%  DE RETENCION",
        "VALOR RETENIDO",
    ])

    sheet = "DETALLE"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        startrow = 6  # espacio para cabecera
        df.to_excel(writer, index=False, sheet_name=sheet, startrow=startrow)
        writer.book.save(out_path)

    wb = load_workbook(out_path)
    ws = wb[sheet]

    ws["A1"] = "DETALLE DE COMPROBANTES DE RETENCION EN LA FUENTE"
    ws["A1"].font = Font(bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws["A1"].alignment = Alignment(horizontal="left")

    ws["A3"] = "RAZON SOCIAL :"
    ws["B3"] = header_razon or ""
    ws["A4"] = "N° RUC :"
    ws["B4"] = header_ruc or ""
    ws["A5"] = "EJERCICIO FISCAL SOLICITADO :"
    ws["B5"] = str(anio)

    # Anchos de columnas
    widths = {1: 26, 2: 45, 3: 24, 4: 20, 5: 16, 6: 16, 7: 16}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Estilos numéricos
    money = NamedStyle(name="money"); money.number_format = '#,##0.00'
    perc  = NamedStyle(name="perc");  perc.number_format = '0.00'
    if "money" not in wb.named_styles: wb.add_named_style(money)
    if "perc"  not in wb.named_styles: wb.add_named_style(perc)

    header_row = 7
    first_data_row = header_row + 1
    last_row = ws.max_row
    for r in range(first_data_row, last_row + 1):
        ws.cell(row=r, column=5).style = "money"  # BASE IMPONIBLE
        ws.cell(row=r, column=7).style = "money"  # VALOR RETENIDO
        # Solo aplicar estilo perc si es número
        val = ws.cell(row=r, column=6).value
        if isinstance(val, (int, float)):
            ws.cell(row=r, column=6).style = "perc"

    wb.save(out_path)

def main():
    anio = input("Ingresa el año (ej. 2022): ").strip()
    # Soporta ejecutar desde la raíz (…/facturas_xml) o desde la carpeta padre (…/).
    # Si BASE_DIR == "." asumimos que estás dentro de facturas_xml/<anio>? entonces year_dir = "."
    if BASE_DIR == Path("."):
        year_dir = Path("./" + anio)
    else:
        year_dir = BASE_DIR / anio

    if not year_dir.exists():
        print(f"No existe: {year_dir}")
        return

    rows = []
    header_razon = ""
    header_ruc = ""

    for mes in MESES:
        mes_dir = year_dir / mes
        if not mes_dir.exists():
            continue
        for xml_file in sorted(glob.glob(str(mes_dir / "*.xml"))):
            h, row = parse_one_xml(Path(xml_file))
            if h.get("razon_suj"): header_razon = h["razon_suj"]
            if h.get("ruc_suj"):   header_ruc   = h["ruc_suj"]
            if row:
                rows.append(row)

    # Orden opcional por fecha y serie
    def key_fecha(r):
        f = r["FECHA DE EMISION DEL COMPROBANTE DE RETENCION"]
        try:
            return datetime.strptime(f, "%Y-%m-%d")
        except:
            return datetime.max
    rows.sort(key=lambda r: (key_fecha(r), r["N° DE COMPROBANTES DE RETENCION"]))

    out_path = Path(OUTPUT.format(anio=anio))
    write_excel(anio, rows, header_razon, header_ruc, out_path)
    print(f"✅ Listo: {out_path.resolve()} (comprobantes: {len(rows)})")

if __name__ == "__main__":
    main()
